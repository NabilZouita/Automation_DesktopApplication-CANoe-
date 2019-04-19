import xlrd
import openpyxl
import os, sys
from xlrd import open_workbook, XLRDError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font 
from Python_CANoe import CANoe


class ExcelPy:
    def __init__(self):
        pass

        #### Test File entered #####
    def TestFile(self,PathExc):
        try:
            self.book = xlrd.open_workbook(PathExc)
        except XLRDError:
            raise RuntimeError("You have entered a wrong File ! Please Check")
        else:
            raise RuntimeError("Thank you!")

    def GetNumbSheets(self,ShPath):
        self.book = xlrd.open_workbook(ShPath)
        self.res = len(self.book.sheet_names())

        return self.res
        
    def ParseFSheet(self,PathExc):

        self.list_Values = []
        self.list_EnVar = []

        self.book = xlrd.open_workbook(PathExc)
        self.sheet = self.book.sheet_by_index(0)
        
        self.List_EnVar = self.sheet.col_values(0)
        self.List_Values = self.sheet.col_values(1)
        
        return self.List_EnVar, self.List_Values 

    def ParseSSheet(self,PathExc,WSheetVal):
        self.LTest_EnVar = []
        self.LTest_Value = []
        
        self.book = xlrd.open_workbook(PathExc)
        self.sheet = self.book.sheet_by_index(WSheetVal)
        
        self.LTest_EnVar = self.sheet.col_values(0)
        self.LTest_Value = self.sheet.col_values(1)
        
        return self.LTest_EnVar, self.LTest_Value
         
         
    def LogEditorPass(self,PathExc,WSheetVal):
        self.wb = openpyxl.load_workbook(PathExc)
        
        self.sheet = self.wb.worksheets[WSheetVal]
        self.sheet.merge_cells('C1:D5')

        self.cell = self.sheet.cell(row=1, column=3)
        self.cell.value = 'Test OK'
        self.cell.alignment = Alignment(horizontal='center', vertical='center')
        self.cell.font = Font(bold=True, size=16, color="008000")

        self.wb.save(PathExc)


    def LogEditorNotPass(self,PathExc,WSheetVal):
        self.wb = openpyxl.load_workbook(PathExc)
        
        self.sheet = self.wb.worksheets[WSheetVal]
        self.sheet.merge_cells('C1:D5')
        
        self.cell = self.sheet.cell(row=1, column=3)
        self.cell.value = 'Test Not OK'
        self.cell.alignment = Alignment(horizontal='center', vertical='center')
        self.cell.font = Font(bold=True, size=16, color="FF6347")

        self.wb.save(PathExc)


class Py_CANoe:
    def __init__(self):
        pass
    # Checking path will be done in Python_CANoe.py
    def FScript_Editor(self,ConfigFile=None,EnvVar=None,EnvVal=None):
        self.file = open("Script.py","w")
        
        self.file.write("from Python_CANoe import CANoe \n")
        self.file.write("var = CANoe() \n")
        self.file.write("var.open_simulation('%s') \n"%ConfigFile)
        self.file.write("var.start_Measurement() \n")        
        for i in range(len(EnvVal)):
            self.file.write("var.set_EnvVar('%s',%s) \n"%(EnvVar[i],EnvVal[i]))


        self.file.close()


    def SScript_Editor(self,envVar=None,valVar=None):
        self.file = open("Script.py","w")
        self.file.write("import time\n")
        self.file.write("from Python_CANoe import CANoe \n")
        self.file.write("var = CANoe() \n")
        self.file.write("time.sleep(1)\n")
        for i in range(len(envVar)):
            self.file.write("var.set_EnvVar('%s',%s) \n"%(envVar[i],valVar[i]))

        self.file.close()        


    def PyCan_Exec(self,path):
        self.root = CANoe()
        self.root.open_simulation(path) #CANoe function restriction are in package
        self.root.start_Measurement()
        self.root.set_EnvVar()

